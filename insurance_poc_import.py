from collections.abc import Iterable, Iterator
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_WORKBOOK = "insurance test for POC dataset.xlsx"
CLEANED_DATASET_SHEET = "cleaned dataset"
CLAIMS_COLLECTION = "insurance_poc_claims"
ENTITY_SCHEMAS_COLLECTION = "insurance_poc_entity_schemas"

SECTION_FIELDS = {
    "claimant": [
        "age",
        "gender",
        "is_senior_citizen",
        "vulnerable_category",
        "ped_type",
    ],
    "policy": [
        "continuous_coverage_months",
        "ped_waiting_period_served",
        "specific_disease_waiting_period_served",
        "grace_period_status",
        "portability_status",
        "sum_insured_enhanced",
        "sum_insured_inr",
        "annual_premium_inr",
        "policy_zone",
    ],
    "hospitalization": [
        "diagnosis_category",
        "treatment_system",
        "is_modern_treatment",
        "hospital_accreditation",
        "hospitalization_type",
        "hospital_zone",
        "admission_type",
        "days_hospitalised",
        "length_of_stay_hours",
        "active_treatment_flag",
        "hospital_risk_score",
    ],
    "financials": [
        "claim_amount_inr",
        "standard_rate_inr",
        "standard_rate_deviation_pct",
        "claim_coverage_ratio",
        "bill_structure",
        "non_medical_expenses_pct",
        "non_medical_deduction_inr",
        "disease_sub_limit_applies",
        "sub_limit_amount_inr",
        "disease_sub_limit_capped",
        "sub_limit_excess_amount_inr",
        "zone_mismatch",
        "zone_copay_amount_inr",
        "mandatory_copay_pct",
        "copay_deduction_inr",
        "settled_amount_inr",
        "settlement_ratio_pct",
    ],
    "fraudRisk": [
        "provider_blacklist_flag",
        "investigation_triggered",
        "investigation_outcome",
        "disclosure_of_material_fact",
        "fraud_indicator",
    ],
    "documents": [
        "pre_auth_status",
        "pre_auth_tat_hours",
        "insurer_cashless_tat_hours",
        "free_look_cancellation",
        "document_submission_mode",
        "document_status",
        "claim_submission_timeline_days",
        "previous_claims_count",
    ],
    "settlement": [
        "claim_settlement_type",
        "claim_status",
        "rejection_reason",
    ],
}


def build_claim_document(
    row: dict[str, Any],
    source_row: int,
    import_batch_id: str,
    ingested_at: str,
) -> dict[str, Any]:
    document = {
        "_id": f"insurance-poc.cleaned-dataset.{source_row:06d}",
        "documentType": "insurance_poc_claim",
        "sourceWorkbook": SOURCE_WORKBOOK,
        "sourceSheet": CLEANED_DATASET_SHEET,
        "sourceRow": source_row,
        "importBatchId": import_batch_id,
        "ingestedAt": ingested_at,
    }

    for section, fields in SECTION_FIELDS.items():
        document[section] = {
            field: _clean_value(row.get(field))
            for field in fields
            if field in row
        }

    document["raw"] = {
        field: _clean_value(value)
        for field, value in row.items()
    }
    return document


def build_cleaned_dataset_entity_schema_document(
    headers: Iterable[str],
    source_workbook: str,
    row_count: int,
    ingested_at: str,
) -> dict[str, Any]:
    header_list = list(headers)
    return {
        "_id": "insurance_poc.cleaned_dataset",
        "documentType": "insurance_poc_entity_schema",
        "entityName": "cleaned_dataset",
        "sourceWorkbook": source_workbook,
        "sourceSheet": CLEANED_DATASET_SHEET,
        "rowCount": row_count,
        "ingestedAt": ingested_at,
        "targetCollection": CLAIMS_COLLECTION,
        "columns": [
            {"name": header, "ordinal": index}
            for index, header in enumerate(header_list, start=1)
        ],
        "documentSections": {
            section: [field for field in fields if field in header_list]
            for section, fields in SECTION_FIELDS.items()
        },
    }


def iter_cleaned_dataset_documents(
    workbook_path: Path,
    import_batch_id: str,
    ingested_at: str | None = None,
) -> Iterator[dict[str, Any]]:
    timestamp = ingested_at or datetime.now(UTC).isoformat()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if CLEANED_DATASET_SHEET not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {CLEANED_DATASET_SHEET}")

        sheet = workbook[CLEANED_DATASET_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return

        headers = [_clean_header(value) for value in header_row]
        for source_row, row_values in enumerate(rows, start=2):
            row = dict(zip(headers, row_values, strict=False))
            yield build_claim_document(row, source_row, import_batch_id, timestamp)
    finally:
        workbook.close()


def get_cleaned_dataset_headers_and_count(workbook_path: Path) -> tuple[list[str], int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if CLEANED_DATASET_SHEET not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {CLEANED_DATASET_SHEET}")

        sheet = workbook[CLEANED_DATASET_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            return [], 0

        headers = [_clean_header(value) for value in header_row]
        return headers, max(sheet.max_row - 1, 0)
    finally:
        workbook.close()


def _clean_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value
